# -*- coding: utf-8 -*-
"""
Created on Mon Dec 11 15:06:06 2023

@author: barto
"""
from casadi import *
import casadi as ca
import pandas as pd
from carica_gen_foto import carica_gen_foto
from update_gen_foto import update_gen_foto
from carica_gen_comb import carica_gen_comb 
from update_gen_comb import update_gen_comb 
from update_gen_comb_ca import update_gen_comb_ca
from calcola_MPP_Ac_Power import calcola_MPP_Ac_Power
from carica_gen_eoli import carica_gen_eoli 
from update_gen_eoli import update_gen_eoli 
from update_gen_eoli_ca1 import update_gen_eoli_ca1 
from carica_carico_L1 import carica_carico_L1
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Specificare il nome del file Excel e il nome del foglio all'interno del file
excel_filename = 'DUMMY_beta_01_eq.xlsx'
sheet_name1 = 'gen_foto'
sheet_name2 = 'gen_comb'
sheet_name3 = 'gen_eoli'
sheet_name4 = 'carico_L1'
sheet_name5 ='Tariffa'
timespan = pd.read_excel(excel_filename, sheet_name='Meteo', usecols='A', skiprows=2, header=None, nrows=800, names=['time'])
N = len(timespan)    #ORIZZONTE DI SIMULAZIONE E NUMERO VARIABILI DI OTTIMIZZAZIONE PER CIASCUN DISPOSITIVO, N=96

#------------------------ INIZIALIZZAZIONE -----------------------------------------------------------------------------------------------------------------------------------

# Carica le condizioni iniziali per ogni dispositivo
x_0 = carica_gen_foto(excel_filename, sheet_name1)
x_0num = carica_gen_foto(excel_filename, sheet_name1) 
y_0 = carica_gen_comb(excel_filename, sheet_name2)
y_0num = carica_gen_comb(excel_filename, sheet_name2)
e_0 = carica_gen_eoli(excel_filename, sheet_name3)
e_0num = carica_gen_eoli(excel_filename, sheet_name3)
z_0 = carica_carico_L1(excel_filename, sheet_name4)
z_C = z_0['C']
#Stampare i valori iniziali dei set-point di ogni dispositivo: 


#------------------------ CREAZIONE VARIABILI SIMBOLICHE -----------------------------------------------------------------------------------------------------------------------------------

#Creazione delle variabili simboliche, occorre specificare le dimensioni delle variabili ovvero N.
alpha_x_sym = SX.sym("alpha_x_sym",N)      #var.sim. per generatore fotovoltaico
alpha_y_sym = SX.sym("alpha_y_sym",N)      #var.sim. per generatore a combustibile
alpha_e_sym = SX.sym("alpha_e_sym",N)      #var.sim. per generatore eolico
alpha_all = vertcat(alpha_x_sym, alpha_y_sym, alpha_e_sym) 
#------------------------ AGGIORNAMENTO -----------------------------------------------------------------------------------------------------------------------------------

#Aggiornamento degli oggetti che diventano simbolici
x_updated = update_gen_foto(x_0, alpha_x_sym)                #aggiornamento per generatore fotovoltaico
y_updated = update_gen_comb_ca(y_0, alpha_y_sym)             #aggiornamento per generatore a combustibile
e_updated = update_gen_eoli_ca1(e_0, alpha_e_sym)            #aggiornamento per generatore eolico

#------------------------ IMPOSTAZIONE PROBLEMA -----------------------------------------------------------------------------------------------------------------------------------

#Potenze dei generatori, dell'accumulatore e del carico che diventano simbolici
x_G_sym = x_updated['W']    #voce dizionario potenza generatore fotovoltaico->simbolica
y_G_sym = y_updated['W']    #voce dizionario potenza generatore combustibile->simbolica
y_cF = y_updated['cF']      #voce dizionario  costo combustibile [euro/kg]
y_F =  y_updated['F']       #voce dizionario fuel rate [kg h−1]
e_G_sym = e_updated['W']    #voce dizionario potenza generatore eolico->simbolica
Z_C_sym=  z_0['W']          #voce dizionario potenza carico L1->simbolica

#CALCOLO DELLA POTENZA EROGATA complessivamente dai dispositivi
W_sym=x_G_sym+e_G_sym+y_G_sym-z_C

#LETTURA E ASSEGNAZIONE DEI VALORI PREZZO DURANTE LA GIORNATA ALLA VARIABILE df_tariffa
df_tariffa = pd.read_excel(excel_filename, sheet_name5, skiprows=3, header=None, nrows=800, usecols=[11, 12])
print(len(df_tariffa))

# Comando che utilizza casadi.if_else per selezionare i valori di C_wf in base al segno di W
# dove cWf è uguale a “-prezzo di vendita” per gli istanti in cui W>0 e a “-prezzo di acquisto” per gli istanti in cui W<0
C_wf = ca.if_else(W_sym > 0, -df_tariffa.iloc[:, 0].values, ca.if_else(W_sym == 0, 0, -df_tariffa.iloc[:, 1].values))
C1=0.01
# Minimizzazione della funzione costo f:
f= sum1((W_sym*C_wf)/100 + y_cF*y_F) + C1*sum1(alpha_all)
g = y_updated['Psi'] 
#------------------------ VINCOLI -----------------------------------------------------------------------------------------------------------------------------------

#VINCOLI VARIABILI OTTIMIZZAZIONE

#vincoli inferiori variabili di ottimizzazione
lbxgf=DM.zeros(N)          #vincolo inferiore generatore fotovoltaico
lbxgc=DM.zeros(N)          #vincolo inferiore generatore a combustibile_BMS 
lbxeo=DM.zeros(N)          #vincolo inferiore generatore eolico

#vincoli inferiori variabili di ottimizzazione TOTALI
lbxtot=vertcat(lbxgf,lbxgc,lbxeo)
print("vincoli inferiori variabili:", lbxtot.shape) #STAMPA LE DIMENSIONI DEI VINCOLI TOTALI DELLE VARIABILI INFERIORI

#vincoli superiori variabili di ottimizzazione
ubxgf=DM.ones(N)          #vincolo superiore generatore fotovoltaico
ubxgc=DM.ones(N)           #vincolo superiore generatore a combustibile_BMS 
ubxeo=DM.ones(N)           #vincolo superiore generatore eolico

#vincoli superiori variabili di ottimizzazione TOTALI
ubxtot=vertcat(ubxgf,ubxgc,ubxeo)
print("vincoli superiori variabili:", ubxtot.shape) #STAMPA LE DIMENSIONI DEI VINCOLI TOTALI DELLE VARIABILI SUPERIORI

# vincoli processo - accensioni generatore elettrico
ubg1=0
lbg1=-100000#.inf(1)

#------------------------ RISOLUTORE -----------------------------------------------------------------------------------------------------------------------------------

nlp = {'x': vertcat(alpha_x_sym, alpha_y_sym, alpha_e_sym), 'f': f, 'g': g}

# Ricava i valori numerici per le variabili iniziali
x0_values = vertcat(x_0num['alpha'],y_0num['alpha'], e_0num['alpha'])

# Creazione del solver
solver = nlpsol('solver', 'ipopt', nlp)

# Risoluzione del problema
res = solver(x0=x0_values, lbx=lbxtot, ubx=ubxtot, lbg=lbg1, ubg=ubg1)

# Estrazione dei risultati, e cioè delle variabili di ottimizzazione per ciascun dispositivo
alpha_x_result = res['x'][:N].full().flatten()  # Estrae i risultati di alpha_x_sym, variabile ottimizzazione generatore fotovoltaico_Sandia 
alpha_y_result = res['x'][N:2*N].full().flatten()  # Estrai i risultati di alpha_y_sym, variabile ottimizzazione generatore combustibile
alpha_e_result = res['x'][2*N:3*N].full().flatten() # Estrai i risultati di alpha_e_sym, variabile ottimizzazione generatore eolico  

#Gli oggetti vengono aggiornati, le variabili di ottimizzazione determinate vengono inserite come argomento.
x_updated = update_gen_foto(x_0num,alpha_x_result)
y_updated = update_gen_comb(y_0num, alpha_y_result)
e_updated = update_gen_eoli(e_0num, alpha_e_result)

print("combustibile", y_updated['F'])  # Consumo combustibile in [kg/h]
timespan['rounded_time'] = timespan['time'].apply(lambda x: (x.hour + x.minute / 60))
#------------------------ GRAFICI -----------------------------------------------------------------------------------------------------------------------------------

#Calcolo di C_wf con i risultati ottenuti
W_value1 = x_updated['W'] + y_updated['W'] + e_updated['W'] - z_C  # Calcolo di W_sym usando i risultati ottenuti
C_wf_value1 = ca.if_else(W_value1 > 0, -df_tariffa.iloc[:, 0].values, ca.if_else(W_value1 == 0, 0, -df_tariffa.iloc[:, 1].values))

print('W_sym_numerico', W_value1)
print('C_wf_numerico', C_wf_value1)
print('W_sym*C_wf_numerico',W_value1*C_wf_value1/100)
print('F*cF_numerico', y_updated['F']*y_updated['cF'])
print("Finale_funzione obiettivo",(W_value1*C_wf_value1/100) + y_updated['F']*y_updated['cF'])

# GRAFICO PROFILO DEI SET-POINT DEI DISPOSITIVI
plt.figure(1)
plt.plot(timespan['rounded_time'], alpha_x_result, marker='o', linestyle='-', label='alpha_x_sym')
plt.plot(timespan['rounded_time'], alpha_y_result, marker='s', linestyle='-', label='alpha_y_sym')
plt.plot(timespan['rounded_time'], alpha_e_result, marker='h', linestyle='-', label='alpha_e_sym')
#plt.plot(timespan['rounded_time'], beta_b_result, marker='p', linestyle='-', label='beta_b_sym')
plt.legend()

# Aggiunta di etichette agli assi del primo grafico
plt.xlabel('Time(hr)')
plt.ylabel('Alpha Values')
plt.title('Device set-point profile as a function of time')

# GRAFICO PROFILO DELLE POTENZE EROGATE
plt.figure(2)
plt.plot(timespan['rounded_time'],x_updated['W'], marker='o', linestyle='-', label='Potenza fotovoltaico')
plt.plot(timespan['rounded_time'],y_updated['W'], marker='s', linestyle='-', label='Potenza generatore combustibile')
plt.plot(timespan['rounded_time'],e_updated['W'], marker='h', linestyle='-', label='Potenza eolico')
plt.plot(timespan['rounded_time'],z_C, marker='s', linestyle='-', label='Potenza carico L1')
plt.legend()

# Aggiunta di etichette agli assi del secondo grafico
plt.xlabel('Time(hr)')
plt.ylabel('Power[kWe]')
plt.title('Power profile as a function of time')


# GRAFICO PROFILO DELLE POTENZE EROGATE
plt.figure(3)
plt.plot(timespan['rounded_time'],y_updated['W'], marker='s', linestyle='-', label='Potenza generatore combustibile')
plt.legend()

# Aggiunta di etichette agli assi del secondo grafico
plt.xlabel('Time(hr)')
plt.ylabel('Power[kWe]')
plt.title('Power profile as a function of time')

plt.figure(4)
plt.plot(timespan['rounded_time'],y_updated['W']+x_updated['W']+e_updated['W']-z_C, marker='s', linestyle='-', label='Potenza risultante')
plt.legend()

# Aggiunta di etichette agli assi del secondo grafico
plt.xlabel('Time(hr)')
plt.ylabel('Power[kWe]')
plt.title('Power profile as a function of time')

excel_file_path = 'DUMMY_beta_01_eq.xlsx'

# Creazione di un DataFrame con i risultati
results_df = pd.DataFrame({
    'Rounded Time': timespan['rounded_time'],
    'Alpha_x_Result': alpha_x_result,
    'Alpha_y_Result': alpha_y_result,
    'Alpha_e_Result': alpha_e_result,
    'X_Updated_W': x_updated['W'],
    'Y_Updated_W': y_updated['W'],
    'E_Updated_W': e_updated['W'],
    'Carico L1': z_C
})

# Apri il file Excel esistente con Openpyxl
try:
    wb = load_workbook(excel_file_path)
except FileNotFoundError:
    # Se il file Excel non esiste, crea un nuovo workbook
    wb = Workbook()

# Seleziona o crea il foglio di lavoro 'Risultati'
sheet_name = 'Risultati'
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(title=sheet_name)

# Aggiungi i dati del DataFrame al foglio di lavoro
for row in dataframe_to_rows(results_df, index=False, header=True):
    ws.append(row)

# Salva il file Excel
wb.save(excel_file_path)


# Aggiunta di una legenda
plt.legend()

# Visualizzazione del grafico
plt.show()






